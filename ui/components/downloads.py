from __future__ import annotations

import io
import json
import pandas as pd
import streamlit as st
from typing import Dict, List
from datetime import datetime

from sem_plan.core.types import CampaignOutputs, KeywordRecord


def _create_search_keywords_csv(outputs: CampaignOutputs) -> tuple[str, bytes]:
    """Create comprehensive search keywords CSV."""
    data = []
    for kw in outputs.search_keywords:
        # Ensure volume is an integer
        volume = kw.gkp_avg_monthly_searches or kw.volume or 0
        if volume is not None:
            try:
                volume = int(volume)
            except (ValueError, TypeError):
                volume = 0
        else:
            volume = 0
        
        cpc_low = kw.gkp_top_of_page_bid_low or 1.0
        cpc_high = kw.gkp_top_of_page_bid_high or 2.0
        
        data.append({
            "Ad Group": kw.cluster,
            "Keyword": kw.keyword,
            "Match Type": kw.match_type,
            "Intent": kw.intent,
            "Competition": kw.competition or "medium",
            "Avg Monthly Searches": volume,
            "CPC Low ($)": cpc_low,
            "CPC High ($)": cpc_high,
            "CPC Range": f"${cpc_low:.2f} - ${cpc_high:.2f}",
            "Sources": ", ".join(kw.sources) if kw.sources else "Unknown",
            "Estimated Monthly Cost": volume * cpc_low if volume and cpc_low else 0,
            "ROI Potential": _calculate_roi_potential(volume, cpc_low, kw.intent)
        })
    
    df = pd.DataFrame(data)
    df = df.sort_values(["Ad Group", "Avg Monthly Searches"], ascending=[True, False])
    
    bio = io.BytesIO()
    df.to_csv(bio, index=False)
    return "search_campaign_keywords.csv", bio.getvalue()


def _create_pmax_themes_csv(outputs: CampaignOutputs) -> tuple[str, bytes]:
    """Create PMax themes CSV."""
    data = []
    for i, theme in enumerate(outputs.pmax_themes):
        data.append({
            "Theme ID": i + 1,
            "Theme": theme,
            "Category": _categorize_pmax_theme(theme),
            "Description": _generate_theme_description(theme)
        })
    
    df = pd.DataFrame(data)
    
    bio = io.BytesIO()
    df.to_csv(bio, index=False)
    return "pmax_themes.csv", bio.getvalue()


def _create_shopping_cpc_csv(outputs: CampaignOutputs) -> tuple[str, bytes]:
    """Create shopping CPC recommendations CSV."""
    data = []
    
    # Calculate shopping metrics
    total_volume = 0
    for kw in outputs.search_keywords:
        volume = kw.gkp_avg_monthly_searches or kw.volume or 0
        if volume is not None:
            try:
                total_volume += int(volume)
            except (ValueError, TypeError):
                continue
    avg_cpc = outputs.shopping_target_cpc
    
    data.append({
        "Metric": "Target CPC",
        "Value": str(f"${avg_cpc:.2f}"),
        "Description": "Recommended target cost-per-click for shopping campaigns"
    })
    
    data.append({
        "Metric": "Total Search Volume",
        "Value": str(f"{total_volume:,}"),
        "Description": "Combined monthly search volume across all keywords"
    })
    
    data.append({
        "Metric": "Estimated Monthly Spend",
        "Value": str(f"${total_volume * avg_cpc:.0f}"),
        "Description": "Estimated monthly spend based on target CPC"
    })
    
    # Add keyword-level recommendations
    for kw in outputs.search_keywords:
        if kw.intent in ["transactional", "commercial"]:
            volume = kw.gkp_avg_monthly_searches or kw.volume or 0
            if volume is not None:
                try:
                    volume = int(volume)
                except (ValueError, TypeError):
                    volume = 0
            else:
                volume = 0
            
            cpc_low = kw.gkp_top_of_page_bid_low or avg_cpc
            
            data.append({
                "Metric": f"Keyword: {kw.keyword}",
                "Value": str(f"${cpc_low:.2f}"),
                "Description": f"Volume: {volume:,}, Intent: {kw.intent}, Competition: {kw.competition}"
            })
    
    df = pd.DataFrame(data)
    
    bio = io.BytesIO()
    df.to_csv(bio, index=False)
    return "shopping_cpc_recommendations.csv", bio.getvalue()


def _create_campaign_summary_json(outputs: CampaignOutputs) -> tuple[str, bytes]:
    """Create comprehensive campaign summary JSON."""
    summary = {
        "generated_at": datetime.now().isoformat(),
        "campaign_overview": {
            "total_keywords": len(outputs.search_keywords),
            "pmax_themes_count": len(outputs.pmax_themes),
            "shopping_target_cpc": outputs.shopping_target_cpc,
            "total_search_volume": sum(
                int(kw.gkp_avg_monthly_searches or kw.volume or 0) 
                for kw in outputs.search_keywords 
                if kw.gkp_avg_monthly_searches is not None or kw.volume is not None
            )
        },
        "ad_groups": {},
        "intent_distribution": {},
        "competition_distribution": {},
        "top_keywords": [],
        "pmax_themes": outputs.pmax_themes,
        "recommendations": _generate_strategic_recommendations(outputs)
    }
    
    # Ad group analysis
    ad_groups = {}
    for kw in outputs.search_keywords:
        ad_group = kw.cluster
        if ad_group not in ad_groups:
            ad_groups[ad_group] = {
                "keywords": [],
                "total_volume": 0,
                "avg_cpc": 0,
                "intents": set()
            }
        
        volume = kw.gkp_avg_monthly_searches or kw.volume or 0
        if volume is not None:
            try:
                volume = int(volume)
            except (ValueError, TypeError):
                volume = 0
        else:
            volume = 0
        
        cpc = kw.gkp_top_of_page_bid_low or 1.0
        
        ad_groups[ad_group]["keywords"].append(kw.keyword)
        ad_groups[ad_group]["total_volume"] += volume
        ad_groups[ad_group]["intents"].add(kw.intent)
    
    # Calculate averages and convert sets to lists
    for ad_group, data in ad_groups.items():
        data["avg_cpc"] = data["total_volume"] / len(data["keywords"]) if data["keywords"] else 0
        data["intents"] = list(data["intents"])
        summary["ad_groups"][ad_group] = data
    
    # Intent distribution
    intent_counts = {}
    for kw in outputs.search_keywords:
        intent = kw.intent
        intent_counts[intent] = intent_counts.get(intent, 0) + 1
    summary["intent_distribution"] = intent_counts
    
    # Competition distribution
    competition_counts = {}
    for kw in outputs.search_keywords:
        competition = kw.competition or "medium"
        competition_counts[competition] = competition_counts.get(competition, 0) + 1
    summary["competition_distribution"] = competition_counts
    
    # Top keywords
    def get_volume(kw):
        volume = kw.gkp_avg_monthly_searches or kw.volume or 0
        if volume is not None:
            try:
                return int(volume)
            except (ValueError, TypeError):
                return 0
        return 0
    
    sorted_keywords = sorted(
        outputs.search_keywords,
        key=get_volume,
        reverse=True
    )
    
    for kw in sorted_keywords[:10]:
        volume = kw.gkp_avg_monthly_searches or kw.volume or 0
        if volume is not None:
            try:
                volume = int(volume)
            except (ValueError, TypeError):
                volume = 0
        else:
            volume = 0
        
        summary["top_keywords"].append({
            "keyword": kw.keyword,
            "ad_group": kw.cluster,
            "search_volume": volume,
            "cpc": kw.gkp_top_of_page_bid_low or 1.0,
            "intent": kw.intent,
            "competition": kw.competition or "medium"
        })
    
    bio = io.BytesIO()
    bio.write(json.dumps(summary, indent=2).encode('utf-8'))
    return "campaign_summary.json", bio.getvalue()


def _create_excel_report(outputs: CampaignOutputs) -> tuple[str, bytes]:
    """Create comprehensive Excel report with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    
    # Search Keywords Sheet
    ws1 = wb.active
    ws1.title = "Search Keywords"
    
    # Headers
    headers = ["Ad Group", "Keyword", "Match Type", "Intent", "Competition", 
               "Avg Monthly Searches", "CPC Low ($)", "CPC High ($)", "ROI Potential"]
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    row = 2
    for kw in outputs.search_keywords:
        volume = kw.gkp_avg_monthly_searches or kw.volume or 0
        if volume is not None:
            try:
                volume = int(volume)
            except (ValueError, TypeError):
                volume = 0
        else:
            volume = 0
        
        cpc_low = kw.gkp_top_of_page_bid_low or 1.0
        cpc_high = kw.gkp_top_of_page_bid_high or 2.0
        
        ws1.cell(row=row, column=1, value=kw.cluster)
        ws1.cell(row=row, column=2, value=kw.keyword)
        ws1.cell(row=row, column=3, value=kw.match_type)
        ws1.cell(row=row, column=4, value=kw.intent)
        ws1.cell(row=row, column=5, value=kw.competition or "medium")
        ws1.cell(row=row, column=6, value=volume)
        ws1.cell(row=row, column=7, value=cpc_low)
        ws1.cell(row=row, column=8, value=cpc_high)
        ws1.cell(row=row, column=9, value=_calculate_roi_potential(volume, cpc_low, kw.intent))
        row += 1
    
    # PMax Themes Sheet
    ws2 = wb.create_sheet("PMax Themes")
    ws2.cell(row=1, column=1, value="Theme").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Category").font = Font(bold=True)
    
    for i, theme in enumerate(outputs.pmax_themes, 2):
        ws2.cell(row=i, column=1, value=theme)
        ws2.cell(row=i, column=2, value=_categorize_pmax_theme(theme))
    
    # Summary Sheet
    ws3 = wb.create_sheet("Summary")
    ws3.cell(row=1, column=1, value="Metric").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="Value").font = Font(bold=True)
    
    # Calculate total search volume with proper type handling
    total_volume = 0
    for kw in outputs.search_keywords:
        volume = kw.gkp_avg_monthly_searches or kw.volume or 0
        if volume is not None:
            try:
                total_volume += int(volume)
            except (ValueError, TypeError):
                continue
    
    summary_data = [
        ("Total Keywords", len(outputs.search_keywords)),
        ("PMax Themes", len(outputs.pmax_themes)),
        ("Shopping Target CPC", f"${outputs.shopping_target_cpc:.2f}"),
        ("Total Search Volume", total_volume)
    ]
    
    for i, (metric, value) in enumerate(summary_data, 2):
        ws3.cell(row=i, column=1, value=metric)
        ws3.cell(row=i, column=2, value=value)
    
    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    return "comprehensive_sem_report.xlsx", bio.getvalue()


def _calculate_roi_potential(volume: int, cpc: float, intent: str) -> str:
    """Calculate ROI potential based on volume, CPC, and intent."""
    if volume == 0 or cpc == 0:
        return "Unknown"
    
    volume_score = min(volume / 1000, 5)
    cpc_efficiency = max(1 - (cpc / 5), 0)
    
    intent_multiplier = {
        "transactional": 1.5,
        "commercial": 1.2,
        "informational": 0.8,
        "navigational": 0.6
    }.get(intent, 1.0)
    
    total_score = (volume_score + cpc_efficiency) * intent_multiplier
    
    if total_score >= 4:
        return "High"
    elif total_score >= 2:
        return "Medium"
    else:
        return "Low"


def _categorize_pmax_theme(theme: str) -> str:
    """Categorize PMax theme based on content."""
    theme_lower = theme.lower()
    
    if any(word in theme_lower for word in ["product", "service", "solution"]):
        return "Product Category"
    elif any(word in theme_lower for word in ["use", "case", "scenario", "workflow"]):
        return "Use Case"
    elif any(word in theme_lower for word in ["professional", "business", "enterprise", "team"]):
        return "Demographic"
    elif any(word in theme_lower for word in ["seasonal", "event", "time", "period"]):
        return "Seasonal/Event"
    else:
        return "General"


def _generate_theme_description(theme: str) -> str:
    """Generate description for PMax theme."""
    category = _categorize_pmax_theme(theme)
    
    descriptions = {
        "Product Category": "Focus on specific product or service offerings",
        "Use Case": "Target specific use cases and scenarios",
        "Demographic": "Target specific audience segments",
        "Seasonal/Event": "Time-based or seasonal campaigns",
        "General": "Broad campaign theme"
    }
    
    return descriptions.get(category, "General campaign theme")


def _generate_strategic_recommendations(outputs: CampaignOutputs) -> List[str]:
    """Generate strategic recommendations based on analysis."""
    recommendations = []
    
    if outputs.search_keywords:
        high_volume_keywords = []
        for kw in outputs.search_keywords:
            volume = kw.gkp_avg_monthly_searches or kw.volume or 0
            if volume is not None:
                try:
                    volume = int(volume)
                    if volume > 1000:
                        high_volume_keywords.append(kw)
                except (ValueError, TypeError):
                    continue
        
        low_competition_keywords = [kw for kw in outputs.search_keywords 
                                  if kw.competition == "low"]
        
        if high_volume_keywords:
            recommendations.append(f"Focus on {len(high_volume_keywords)} high-volume keywords for immediate impact")
        
        if low_competition_keywords:
            recommendations.append(f"Target {len(low_competition_keywords)} low-competition keywords for cost efficiency")
        
        if len(outputs.pmax_themes) > 0:
            recommendations.append(f"Implement {len(outputs.pmax_themes)} PMax themes for broad reach")
    
    return recommendations


def render_download_section(outputs: CampaignOutputs) -> None:
    """Render comprehensive download section with multiple export options."""
    
    st.subheader("📥 Export Results")
    
    # File format selection
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**📊 Individual Files**")
        
        # Search Keywords CSV
        search_name, search_data = _create_search_keywords_csv(outputs)
        st.download_button(
            "📋 Search Keywords (CSV)",
            data=search_data,
            file_name=search_name,
            mime="text/csv",
            help="Complete search campaign keywords with ad groups and metrics"
        )
        
        # PMax Themes CSV
        pmax_name, pmax_data = _create_pmax_themes_csv(outputs)
        st.download_button(
            "🎯 PMax Themes (CSV)",
            data=pmax_data,
            file_name=pmax_name,
            mime="text/csv",
            help="Performance Max campaign themes and categories"
        )
        
        # Shopping CPC CSV
        shopping_name, shopping_data = _create_shopping_cpc_csv(outputs)
        st.download_button(
            "🛒 Shopping CPC (CSV)",
            data=shopping_data,
            file_name=shopping_name,
            mime="text/csv",
            help="Shopping campaign CPC recommendations and analysis"
        )
    
    with col2:
        st.write("**📄 Comprehensive Reports**")
        
        # Campaign Summary JSON
        summary_name, summary_data = _create_campaign_summary_json(outputs)
        st.download_button(
            "📈 Campaign Summary (JSON)",
            data=summary_data,
            file_name=summary_name,
            mime="application/json",
            help="Complete campaign analysis and recommendations"
        )
        
        # Excel Report
        try:
            excel_name, excel_data = _create_excel_report(outputs)
            st.download_button(
                "📊 Complete Report (Excel)",
                data=excel_data,
                file_name=excel_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Comprehensive Excel report with multiple sheets"
            )
        except ImportError:
            st.info("💡 Install openpyxl for Excel export: `pip install openpyxl`")
    
    # Export options info
    with st.expander("💡 Export Options Guide", expanded=False):
        st.write("""
        **Available Export Formats:**
        
        📋 **CSV Files** - Individual data files for specific campaign types
        - Search Keywords: Complete keyword list with ad groups and metrics
        - PMax Themes: Performance Max campaign themes and categories  
        - Shopping CPC: Shopping campaign recommendations and analysis
        
        📄 **Comprehensive Reports**
        - JSON Summary: Complete campaign analysis with strategic insights
        - Excel Report: Multi-sheet report with all data and visualizations
        
        **Usage Tips:**
        - Use CSV files for direct import into Google Ads
        - Use JSON summary for API integration or further analysis
        - Use Excel report for presentations and detailed analysis
        """)
